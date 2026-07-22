"""Modular customer export engine.

Supports CSV and JSON today. New formats can be added later by writing
a function with the signature (rows, path) -> None and registering it
in EXPORTERS; no other code needs to change.
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

from config import EXPORTS_DIR, ensure_directories


EXPORT_FIELDS = [
    "Loan Number",
    "First Name",
    "Last Name",
    "Phone Numbers",
    "Balance",
    "Days Overdue",
    "Monthly Payment",
    "Current Overdue Amount",
    "Original Loan Amount",
    "Final Status",
    "Status Timestamp",
    "Last Edited Timestamp",
    "Session ID",
]

SUPPORTED_FORMATS = ("csv", "json")


class ExportError(Exception):
    """Raised when an export cannot be completed."""


def _customer_row(customer: dict[str, Any], session_id: Any) -> dict[str, Any]:
    return {
        "Loan Number": customer["loan_number"],
        "First Name": customer["first_name"],
        "Last Name": customer["last_name"],
        "Phone Numbers": ", ".join(customer.get("phone_numbers") or []),
        "Balance": customer.get("balance", ""),
        "Days Overdue": customer.get("days_overdue", ""),
        "Monthly Payment": customer.get("monthly_payment", ""),
        "Current Overdue Amount": customer.get("current_overdue_amount", ""),
        "Original Loan Amount": customer.get("original_loan_amount", ""),
        "Final Status": customer.get("status", ""),
        "Status Timestamp": customer.get("status_timestamp") or "",
        "Last Edited Timestamp": customer.get("last_edited_timestamp") or "",
        "Session ID": "" if session_id is None else str(session_id),
    }


def _export_csv(rows: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def _export_json(rows: list[dict[str, Any]], path: Path) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def _export_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    """Export customers to an Excel (.xlsx) workbook."""
    try:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill  # noqa: PLC0415
    except ImportError as exc:
        raise ExportError(
            "openpyxl is required for Excel export. Run: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    # Header row with bold styling and a light-blue fill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.append(EXPORT_FIELDS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row in rows:
        ws.append([row.get(field, "") for field in EXPORT_FIELDS])

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(path)


# Registry of available export formats. Add new formats here (e.g. Excel,
# PDF) by mapping a format name to a (rows, path) -> None function.
EXPORTERS: dict[str, Callable[[list[dict[str, Any]], Path], None]] = {
    "csv": _export_csv,
    "json": _export_json,
    "xlsx": _export_xlsx,
}


def export_customers(
    customers: list[dict[str, Any]],
    session_id: Any,
    export_format: str,
) -> Path:
    """Write customers to a timestamped file in EXPORTS_DIR and return its path."""
    export_format = export_format.lower().strip()
    exporter = EXPORTERS.get(export_format)
    if exporter is None:
        raise ExportError(
            f"Unsupported export format: {export_format}. "
            f"Supported formats: {', '.join(SUPPORTED_FORMATS)}."
        )
    if not customers:
        raise ExportError("No customers to export.")

    ensure_directories()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = EXPORTS_DIR / f"{timestamp}_export.{export_format}"
    rows = [_customer_row(customer, session_id) for customer in customers]

    try:
        exporter(rows, path)
    except OSError as exc:
        raise ExportError("Could not write the export file.") from exc

    return path
