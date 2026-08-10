from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from ai_parser import AIParser
from config import Settings
from database import Database
from importer import Importer, ImporterError
from telegram_ui import PROGRESS_STAGES, run_progress
from validation import ValidationError, load_json_array, validate_customers


class FakeParser(AIParser):
    def __init__(self):
        self.settings = Settings(telegram_bot_token="test", openai_api_key=None)

    async def parse_text(self, text):
        return [
            {
                "loan_number": "LN-1",
                "first_name": "Ada",
                "last_name": "Lovelace",
                "phone_numbers": ["+1 (555) 101-2020"],
                "balance": "100",
                "days_overdue": "4",
            }
        ]

    async def parse_image(self, image_path: Path):
        return [
            {
                "loan_number": "IMG-1",
                "first_name": "Grace",
                "last_name": "Hopper",
                "phone_numbers": ["555-3030"],
                "balance": "200",
                "days_overdue": "9",
            }
        ]


class FakeMessage:
    def __init__(self):
        self.edits = []

    async def edit_text(self, text):
        self.edits.append(text)


@pytest.fixture()
def importer(tmp_path, monkeypatch):
    import importer as importer_module

    originals = tmp_path / "originals"
    imports = tmp_path / "imports"
    originals.mkdir()
    imports.mkdir()
    monkeypatch.setattr(importer_module, "ORIGINALS_DIR", originals)
    monkeypatch.setattr(importer_module, "IMPORTS_DIR", imports)

    database = Database(tmp_path / "session.db")
    return Importer(FakeParser(), database)


@pytest.mark.asyncio
async def test_import_image_stores_json_customer_and_original(importer, tmp_path):
    image = tmp_path / "crm.jpg"
    image.write_bytes(b"fake-image")

    result = await importer.import_image(image)

    assert result.imported_count == 1
    assert result.customers[0]["loan_number"] == "IMG-1"
    assert result.original_path.exists()
    assert result.normalized_path.exists()


@pytest.mark.asyncio
async def test_import_xlsx_stores_json_customer_and_original(importer, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    sheet.append(["Loan Number", "First Name", "Last Name", "Phone"])
    sheet.append(["XLSX-1", "Hedy", "Lamarr", "555-1234"])
    sheet.append(["XLSX-2", "Ada", "Lovelace", "555-5678"])
    xlsx_path = tmp_path / "customers.xlsx"
    workbook.save(xlsx_path)

    result = await importer.import_xlsx(xlsx_path)

    assert result.imported_count == 2
    assert result.customers[0]["loan_number"] == "XLSX-1"
    assert result.customers[1]["loan_number"] == "XLSX-2"
    assert importer.database.count_customers() == 2
    assert result.original_path.exists()
    assert result.normalized_path.exists()
    assert result.original_path.name == "customers.xlsx"
    assert result.normalized_path.name == "customers.json"


@pytest.mark.asyncio
async def test_import_xlsx_with_no_loan_numbers_is_rejected(importer, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    # "Loan Number" column is missing, so all rows will be rejected.
    sheet.append(["First Name", "Last Name", "Phone"])
    sheet.append(["Hedy", "Lamarr", "555-1234"])
    xlsx_path = tmp_path / "invalid_customers.xlsx"
    workbook.save(xlsx_path)

    with pytest.raises(ImporterError, match="Import is empty: no valid customers found"):
        await importer.import_xlsx(xlsx_path)


@pytest.mark.asyncio
async def test_import_xlsx_with_no_valid_headers_raises_error(importer, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    # None of these headers are recognizable.
    sheet.append(["Column A", "Column B", "Column C"])
    sheet.append(["Hedy", "Lamarr", "555-1234"])
    xlsx_path = tmp_path / "invalid_headers.xlsx"
    workbook.save(xlsx_path)

    with pytest.raises(ImporterError, match="Could not find any recognizable column headers"):
        await importer.import_xlsx(xlsx_path)


@pytest.mark.asyncio
async def test_paste_text_parses_and_updates_database(importer):
    result = await importer.import_text("Ada Lovelace loan LN-1 phone +1 555 101 2020")

    assert result.imported_count == 1
    assert importer.database.count_customers() == 1


@pytest.mark.asyncio
async def test_paste_json_validation_passes(importer):
    payload = """
    [
      {
        "loan_number": "JSON-1",
        "first_name": "Katherine",
        "last_name": "Johnson",
        "phone_numbers": ["555-4040"],
        "balance": "500",
        "days_overdue": "2"
      }
    ]
    """

    result = await importer.import_text(payload)

    assert result.imported_count == 1
    assert result.customers[0]["phone_numbers"] == ["5554040"]


def test_duplicate_loan_numbers_removed():
    result = validate_customers(
        [
            {
                "loan_number": "DUP-1",
                "first_name": "A",
                "last_name": "B",
                "phone_numbers": ["111"],
                "balance": "",
                "days_overdue": "",
            },
            {
                "loan_number": "DUP-1",
                "first_name": "C",
                "last_name": "D",
                "phone_numbers": ["222"],
                "balance": "",
                "days_overdue": "",
            },
        ]
    )

    assert result.ok
    assert len(result.customers) == 1


def test_invalid_json_is_graceful():
    with pytest.raises(ValidationError):
        load_json_array("[not json")


def test_missing_phone_number_fails_validation():
    result = validate_customers(
        [
            {
                "loan_number": "NO-PHONE",
                "first_name": "A",
                "last_name": "B",
                "phone_numbers": [],
                "balance": "",
                "days_overdue": "",
            }
        ]
    )

    assert result.ok
    assert len(result.flagged) == 1
    assert result.flagged[0]["_status"] == "needs_review"
    assert "phone number" in result.flagged[0]["_issue"]


@pytest.mark.asyncio
async def test_empty_import_rejected(importer):
    with pytest.raises(ImporterError):
        await importer.import_text("[]")


@pytest.mark.asyncio
async def test_progress_updates_every_stage():
    message = FakeMessage()

    await run_progress(message)

    assert message.edits == [f"⏳ {stage}" for stage in PROGRESS_STAGES]
